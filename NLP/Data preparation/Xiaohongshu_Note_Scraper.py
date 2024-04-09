from DrissionPage import ChromiumPage
import pandas as pd
from DrissionPage.errors import ElementNotFoundError
from tqdm import tqdm
import time
from urllib.parse import quote
import random
import re
import openpyxl


def sign_in():
    sign_in_page = ChromiumPage()
    sign_in_page.get('https://www.xiaohongshu.com')
    print("Please scan code to log in")
    # The first run requires scanning code login
    time.sleep(20)


def search(keyword):
    global page
    page = ChromiumPage()
    page.get(f'https://www.xiaohongshu.com/search_result?keyword={keyword}&source=web_search_result_notes')


import time

def select_category(category):
    time.sleep(5)  # wait for 5 seconds
    try:
        content_container = page.ele('.content-container')
        videos = content_container.ele(f'text={category}')
        videos.click()
    except ElementNotFoundError:
        print(f"Failed to find element with text '{category}'")



def get_info():
    print(f"第{i}次爬取")
    # Locate sections that contain note information
    container = page.ele('.feeds-page')
    sections = container.eles('.note-item')
    # Extract note information
    for section in sections:
        try:
            # plog link
            note_link = section.ele('tag:a', timeout=0).link
            # Title, author, like
            footer = section.ele('.footer', timeout=0)
            # title
            try:
                title = footer.ele('.title', timeout=0).text
            except:
                title = "blank title"

            # author
            author_wrapper = footer.ele('.author-wrapper')
            author = author_wrapper.ele('.author').text
            author_link = author_wrapper.ele('tag:a', timeout=0).link
            author_img = author_wrapper.ele('tag:img', timeout=0).link

            # like
            like = footer.ele('.like-wrapper like-active').text
            # processing ends with a w
            if like.endswith("w"):
                # Find numbers using regular expressions
                numbers = re.findall(r'\d+\.\d+|\d+', like)
                # Converts a string of numbers to a floating point number
                float_value = float(numbers[0])
                # Multiply a floating-point number by 10,000 to represent 10,000
                like = int(float_value * 10000)
            else:
                like = int(like)

            print(title, author, like, note_link, author_link, author_img)
            contents.append([title, author, note_link, author_link, author_img, like])

        except:
            pass
    print(f"Crawl {i} times, total fetch {len(contents)} bar")


def page_scroll_down():
    print("*******Slide page*********")
    # Generate a random time
    random_time = random.uniform(0.5, 1.5)
    # pause
    time.sleep(random_time)
    # time.sleep(1)
    # page.scroll.down(5000)
    page.scroll.to_bottom()


def crawler(times):
    global i
    for i in tqdm(range(1, times + 1)):
        get_info()
        page_scroll_down()


def auto_resize_column(excel_path):
    """ Adaptive Column width """
    wb = openpyxl.load_workbook(excel_path)
    worksheet = wb.active
    # Loop through columns 1-2 of the worksheet
    for col in worksheet.iter_cols(min_col=1, max_col=2):
        max_length = 0
        # Column name
        column = col[0].column_letter
        # Loop through all the cells in the column
        for cell in col:
            try:
                # If the length of the current cell value is greater than max_length, the value of max_length is updated
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Calculates the adjusted column width
        adjusted_width = (max_length + 2) * 2
        # Use the worksheet.column_dimensions property to set the column width
        worksheet.column_dimensions[column].width = adjusted_width

        # Loop through columns 3-5 in the worksheet
        for col in worksheet.iter_cols(min_col=3, max_col=5):
            max_length = 0
            column = col[0].column_letter  # Get the column name

            # Use the worksheet.column_dimensions property to set the column width
            worksheet.column_dimensions[column].width = 15

    wb.save(excel_path)


def save_to_excel(data):
    # Save to excel file
    name = ['标题', '作者', '笔记链接', '作者主页', '作者头像', '点赞数']
    df = pd.DataFrame(columns=name, data=data)

    # # 写入原文件前清除openpyxl不支持的字符
    # for col in df[['title', 'author']]:
    #     df[col] = df[col].apply(lambda x: ILLEGAL_CHARACTERS_RE.sub(r'', str(x) if not pd.isna(x) else ''))

    df['点赞数'] = df['点赞数'].astype(int)
    # Delete duplicate lines
    df = df.drop_duplicates()
    # Sort in descending order of likes
    df = df.sort_values(by='点赞数', ascending=False)
    # file path
    # global excel_path
    excel_path = f"小红书搜索结果-{category}-{keyword}-{df.shape[0]}条.xlsx"
    df.to_excel(excel_path, index=False)
    print(f"总计向下翻页{times}次，获取到{len(data)}条，去重后剩余{df.shape[0]}条")
    print(f"数据已保存到：{excel_path}")

    # Automatically adjusts excel table column width
    auto_resize_column(excel_path)
    print("An excel table column width has been automatically adjusted")
    # Save note link to txt file
    save_to_txt(data, 'note_links.txt')

def save_to_txt(data, filename):
    with open(filename, 'w') as f:
        for item in data:
            f.write("%s\n" % item[2])  # item[2]Note link


if __name__ == '__main__':
    # The contents list is used to store all the crawled information
    contents = []

    # 1、Set search keywords
    keyword = "How to treat plant rust"
    # 2、Set the number of page crawl down
    times = 20
    # 3、Settings Select the note type. You can select All, Text, Video. The default value is All.
    category = "全部"

    # You need to log in for the first run. You can comment it out
    # sign_in()

    # Keywords are converted to url encoding
    keyword_temp_code = quote(keyword.encode('utf-8'))
    keyword_encode = quote(keyword_temp_code.encode('gb2312'))

    # Search Xiaohongshu articles according to keywords
    search(keyword_encode)

    select_category(category)

    # According to the number of times set, start to crawl data
    crawler(times)

    # The crawled data is saved to a local excel file
    save_to_excel(contents)
